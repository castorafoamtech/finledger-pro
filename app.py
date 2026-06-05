"""
FinLedger Pro — app.py
Ultra-fast, Excel-grade ledger with live KPI dashboard.
Deploy: Render + Neon PostgreSQL | Local: python app.py
"""

import os, json, io
from datetime import date, datetime
from collections import defaultdict

from flask import Flask, render_template, request, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func, or_, text
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
DB_PATH  = os.path.join(BASE_DIR, 'data', 'fin_accounts.db')
os.makedirs(os.path.join(BASE_DIR, 'data'), exist_ok=True)

try:
    from dotenv import load_dotenv
    load_dotenv(os.path.join(BASE_DIR, '.env'))
except ImportError:
    pass

_db_url = os.environ.get('FIN_DATABASE_URL') or os.environ.get('DATABASE_URL') or f'sqlite:///{DB_PATH}'
# Render / Heroku use postgres:// which SQLAlchemy 2.x requires as postgresql+psycopg2://
if _db_url and _db_url.startswith('postgres://'):
    _db_url = 'postgresql+psycopg2://' + _db_url[len('postgres://'):]

app = Flask(__name__, template_folder='templates', static_folder='static')
app.config.update(
    SQLALCHEMY_DATABASE_URI=_db_url,
    SQLALCHEMY_TRACK_MODIFICATIONS=False,
    SQLALCHEMY_ENGINE_OPTIONS={'pool_pre_ping': True, 'pool_recycle': 300},
    SECRET_KEY=os.environ.get('SECRET_KEY', 'fin-master-secret-2026'),
    JSON_SORT_KEYS=False,
)
db = SQLAlchemy(app)


# ═══════════════════════════════════════════════════════════════
# MODELS
# ═══════════════════════════════════════════════════════════════

class FinAccount(db.Model):
    __tablename__ = 'fin_accounts'
    id          = db.Column(db.Integer, primary_key=True)
    name        = db.Column(db.String(200), nullable=False)
    number      = db.Column(db.String(100), default='')
    # group_type: 'H' = H account, 'dolph' = Dolph account, 'regular'
    group_type  = db.Column(db.String(20), default='regular')
    is_pinned   = db.Column(db.Boolean, default=False)
    is_archived = db.Column(db.Boolean, default=False)
    sort_order  = db.Column(db.Integer, default=0)
    color       = db.Column(db.String(20), default='')
    notes       = db.Column(db.Text, default='')
    created_at  = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at  = db.Column(db.DateTime, default=datetime.utcnow)

    transactions = db.relationship(
        'FinTransaction', backref='account',
        lazy='dynamic', cascade='all, delete-orphan'
    )

    def get_totals(self):
        row = db.session.query(
            func.coalesce(func.sum(FinTransaction.credit), 0),
            func.coalesce(func.sum(FinTransaction.debit), 0)
        ).filter_by(account_id=self.id).one()
        return float(row[0]), float(row[1])

    def to_dict(self, with_balance=True):
        d = dict(id=self.id, name=self.name, number=self.number,
                 group_type=self.group_type, is_pinned=self.is_pinned,
                 is_archived=self.is_archived, sort_order=self.sort_order,
                 color=self.color, notes=self.notes,
                 created_at=self.created_at.isoformat() if self.created_at else None)
        if with_balance:
            tc, td = self.get_totals()
            d.update(total_credits=tc, total_debits=td, balance=tc - td)
        return d


class FinTransaction(db.Model):
    __tablename__ = 'fin_transactions'
    id               = db.Column(db.Integer, primary_key=True)
    account_id       = db.Column(db.Integer, db.ForeignKey('fin_accounts.id'), nullable=False)
    txn_date         = db.Column(db.Date, nullable=True)   # actual date entry was recorded
    transaction_date = db.Column(db.Date, nullable=True)   # date the transaction pertains to
    credit        = db.Column(db.Float, default=0)
    debit         = db.Column(db.Float, default=0)
    credit_remark = db.Column(db.String(500), default='')
    debit_remark  = db.Column(db.String(500), default='')
    # status: normal | unpaid | flagged | done | reminder
    status        = db.Column(db.String(20), default='normal')
    # color_flag: '' | red | orange | yellow | green | blue | purple
    color_flag    = db.Column(db.String(20), default='')
    tags          = db.Column(db.String(300), default='')
    priority      = db.Column(db.Integer, default=0)   # 0=normal 1=high 2=urgent
    notes         = db.Column(db.Text, default='')
    sort_order    = db.Column(db.Integer, default=0)
    created_at    = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at    = db.Column(db.DateTime, default=datetime.utcnow)

    def to_dict(self):
        return dict(
            id=self.id, account_id=self.account_id,
            txn_date=self.txn_date.isoformat() if self.txn_date else None,
            transaction_date=self.transaction_date.isoformat() if self.transaction_date else None,
            credit=self.credit, debit=self.debit,
            credit_remark=self.credit_remark, debit_remark=self.debit_remark,
            status=self.status, color_flag=self.color_flag,
            tags=self.tags, priority=self.priority, notes=self.notes,
            sort_order=self.sort_order,
            created_at=self.created_at.isoformat() if self.created_at else None,
            updated_at=self.updated_at.isoformat() if self.updated_at else None,
        )


class FinSetting(db.Model):
    __tablename__ = 'fin_settings'
    key   = db.Column(db.String(80), primary_key=True)
    value = db.Column(db.Text, default='')


class FinAuditLog(db.Model):
    __tablename__ = 'fin_audit_log'
    id          = db.Column(db.Integer, primary_key=True)
    entity_type = db.Column(db.String(50))
    entity_id   = db.Column(db.Integer, nullable=True)
    action      = db.Column(db.String(20))
    before_data = db.Column(db.Text, default='')
    after_data  = db.Column(db.Text, default='')
    description = db.Column(db.String(500), default='')
    created_at  = db.Column(db.DateTime, default=datetime.utcnow)


# ═══════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════

def _get(key, default=''):
    s = FinSetting.query.get(key)
    return s.value if s else default

def _set(key, value):
    s = FinSetting.query.get(key)
    if s:
        s.value = str(value)
    else:
        db.session.add(FinSetting(key=key, value=str(value)))

def _audit(entity_type, entity_id, action, before=None, after=None, desc=''):
    db.session.add(FinAuditLog(
        entity_type=entity_type, entity_id=entity_id, action=action,
        before_data=json.dumps(before) if before else '',
        after_data=json.dumps(after) if after else '',
        description=desc,
    ))

def _parse_date(s):
    if not s:
        return None
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(str(s).strip(), fmt).date()
        except Exception:
            pass
    return None

def compute_kpis():
    """Core KPI calculation — the 3 hero values."""
    t_value        = float(_get('t_value', '0') or 0)
    previous_value = _get('previous_value', '')
    updated_till   = _get('updated_till', '')

    # H accounts
    h_accts = FinAccount.query.filter_by(group_type='H', is_archived=False).all()
    h_balance = sum(a.get_totals()[0] - a.get_totals()[1] for a in h_accts)

    # Dolph accounts (still computed for reference, not included in manual balance)
    d_accts = FinAccount.query.filter_by(group_type='dolph', is_archived=False).all()
    dolph_balance = sum(a.get_totals()[0] - a.get_totals()[1] for a in d_accts)

    # Manual Balance = H + T  (Dolph excluded)
    manual_balance = h_balance + t_value

    # System Total = all credits − all debits (non-archived, excludes H accounts)
    row = db.session.query(
        func.coalesce(func.sum(FinTransaction.credit), 0),
        func.coalesce(func.sum(FinTransaction.debit), 0)
    ).join(FinAccount, FinTransaction.account_id == FinAccount.id)\
     .filter(FinAccount.is_archived == False, FinAccount.group_type != 'H').one()
    system_total = float(row[0]) - float(row[1])

    return dict(
        t_value=t_value,
        previous_value=previous_value,
        updated_till=updated_till,
        h_balance=h_balance,
        dolph_balance=dolph_balance,
        manual_balance=manual_balance,
        system_total=system_total,
        difference=manual_balance - system_total,
        total_credits=float(row[0]),
        total_debits=float(row[1]),
    )


def _acct_dict_with_balance(account):
    tc, td = account.get_totals()
    d = account.to_dict(with_balance=False)
    d.update(total_credits=tc, total_debits=td, balance=tc - td,
             txn_count=account.transactions.count())
    return d


def _max_sort(model_class, **filters):
    q = db.session.query(func.max(model_class.sort_order))
    for k, v in filters.items():
        q = q.filter(getattr(model_class, k) == v)
    return (q.scalar() or 0)


# ═══════════════════════════════════════════════════════════════
# ROUTES — Pages
# ═══════════════════════════════════════════════════════════════

@app.route('/')
def dashboard():
    return render_template('fin_dashboard.html')


# ═══════════════════════════════════════════════════════════════
# ROUTES — KPIs & Settings
# ═══════════════════════════════════════════════════════════════

@app.route('/api/kpis')
def api_kpis():
    return jsonify(compute_kpis())


@app.route('/api/settings', methods=['GET'])
def api_settings_get():
    return jsonify({s.key: s.value for s in FinSetting.query.all()})


@app.route('/api/settings', methods=['PUT'])
def api_settings_put():
    data = request.get_json() or {}
    for k, v in data.items():
        _set(k, v)
    _audit('setting', None, 'update', desc=f'Updated: {list(data.keys())}')
    db.session.commit()
    return jsonify({'ok': True, 'kpis': compute_kpis()})


# ═══════════════════════════════════════════════════════════════
# ROUTES — Accounts
# ═══════════════════════════════════════════════════════════════

@app.route('/api/accounts', methods=['GET'])
def api_accounts_list():
    archived = request.args.get('archived', 'false').lower() == 'true'
    q = FinAccount.query
    if not archived:
        q = q.filter_by(is_archived=False)
    accounts = q.order_by(
        FinAccount.is_pinned.desc(), FinAccount.sort_order, FinAccount.id
    ).all()

    # Batch-load all balances in 2 queries
    credits = dict(
        db.session.query(FinTransaction.account_id, func.sum(FinTransaction.credit))
        .group_by(FinTransaction.account_id).all()
    )
    debits = dict(
        db.session.query(FinTransaction.account_id, func.sum(FinTransaction.debit))
        .group_by(FinTransaction.account_id).all()
    )
    counts = dict(
        db.session.query(FinTransaction.account_id, func.count(FinTransaction.id))
        .group_by(FinTransaction.account_id).all()
    )

    result = []
    for a in accounts:
        tc = float(credits.get(a.id, 0) or 0)
        td = float(debits.get(a.id, 0) or 0)
        d = a.to_dict(with_balance=False)
        d.update(total_credits=tc, total_debits=td, balance=tc - td,
                 txn_count=int(counts.get(a.id, 0)))
        result.append(d)
    return jsonify(result)


@app.route('/api/accounts', methods=['POST'])
def api_accounts_create():
    data = request.get_json() or {}
    acct = FinAccount(
        name=data.get('name', '').strip(),
        number=data.get('number', '').strip(),
        group_type=data.get('group_type', 'regular'),
        color=data.get('color', ''),
        notes=data.get('notes', ''),
        sort_order=_max_sort(FinAccount) + 1,
    )
    db.session.add(acct)
    db.session.flush()
    _audit('account', acct.id, 'create', after=acct.to_dict(), desc=f'Created: {acct.name}')
    db.session.commit()
    return jsonify(_acct_dict_with_balance(acct)), 201


@app.route('/api/accounts/<int:aid>', methods=['GET'])
def api_account_get(aid):
    a = FinAccount.query.get_or_404(aid)
    return jsonify(_acct_dict_with_balance(a))


@app.route('/api/accounts/<int:aid>', methods=['PUT'])
def api_accounts_update(aid):
    a = FinAccount.query.get_or_404(aid)
    before = a.to_dict()
    data = request.get_json() or {}
    for field in ('name', 'number', 'group_type', 'color', 'notes',
                  'is_pinned', 'is_archived', 'sort_order'):
        if field in data:
            setattr(a, field, data[field])
    a.updated_at = datetime.utcnow()
    _audit('account', a.id, 'update', before=before, after=a.to_dict())
    db.session.commit()
    return jsonify(_acct_dict_with_balance(a))


@app.route('/api/accounts/<int:aid>', methods=['DELETE'])
def api_accounts_delete(aid):
    a = FinAccount.query.get_or_404(aid)
    force = request.args.get('force', 'false').lower() == 'true'
    if force:
        _audit('account', aid, 'delete', before=a.to_dict(), desc=f'Deleted: {a.name}')
        db.session.delete(a)
    else:
        _audit('account', aid, 'archive', desc=f'Archived: {a.name}')
        a.is_archived = True
        a.updated_at = datetime.utcnow()
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/api/accounts/reorder', methods=['POST'])
def api_accounts_reorder():
    for item in (request.get_json() or []):
        FinAccount.query.filter_by(id=item['id']).update({'sort_order': item['sort_order']})
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/api/accounts/merge', methods=['POST'])
def api_accounts_merge():
    data = request.get_json() or {}
    src = FinAccount.query.get_or_404(data['source_id'])
    tgt = FinAccount.query.get_or_404(data['target_id'])
    FinTransaction.query.filter_by(account_id=src.id).update({'account_id': tgt.id})
    _audit('account', src.id, 'merge', desc=f'Merged {src.name} → {tgt.name}')
    src.is_archived = True
    db.session.commit()
    return jsonify({'ok': True, 'target': _acct_dict_with_balance(tgt)})


@app.route('/api/accounts/duplicate/<int:aid>', methods=['POST'])
def api_accounts_duplicate(aid):
    src = FinAccount.query.get_or_404(aid)
    copy = FinAccount(
        name=f'{src.name} (Copy)', number=src.number,
        group_type=src.group_type, color=src.color,
        sort_order=_max_sort(FinAccount) + 1,
    )
    db.session.add(copy)
    db.session.flush()
    for t in FinTransaction.query.filter_by(account_id=src.id).all():
        db.session.add(FinTransaction(
            account_id=copy.id, txn_date=t.txn_date,
            credit=t.credit, debit=t.debit,
            credit_remark=t.credit_remark, debit_remark=t.debit_remark,
            status=t.status, color_flag=t.color_flag, sort_order=t.sort_order,
        ))
    db.session.commit()
    return jsonify(_acct_dict_with_balance(copy)), 201


# ═══════════════════════════════════════════════════════════════
# ROUTES — Ledger / Transactions
# ═══════════════════════════════════════════════════════════════

@app.route('/api/accounts/<int:aid>/ledger', methods=['GET'])
def api_ledger(aid):
    a = FinAccount.query.get_or_404(aid)
    args = request.args

    q = FinTransaction.query.filter_by(account_id=aid)

    txn_type = args.get('type', 'all')
    if txn_type == 'credits':
        q = q.filter(FinTransaction.credit > 0)
    elif txn_type == 'debits':
        q = q.filter(FinTransaction.debit > 0)

    if args.get('status'):
        q = q.filter(FinTransaction.status == args['status'])
    if args.get('color_flag'):
        q = q.filter(FinTransaction.color_flag == args['color_flag'])

    d_from = _parse_date(args.get('date_from'))
    d_to   = _parse_date(args.get('date_to'))
    if d_from:
        q = q.filter(FinTransaction.txn_date >= d_from)
    if d_to:
        q = q.filter(FinTransaction.txn_date <= d_to)

    if args.get('q'):
        p = f'%{args["q"]}%'
        q = q.filter(or_(
            FinTransaction.credit_remark.ilike(p),
            FinTransaction.debit_remark.ilike(p),
            FinTransaction.notes.ilike(p),
            FinTransaction.tags.ilike(p),
        ))

    txns = q.order_by(FinTransaction.sort_order, FinTransaction.id).all()

    running = 0
    rows, fc, fd = [], 0.0, 0.0
    for t in txns:
        running += t.credit - t.debit
        fc += t.credit
        fd += t.debit
        d = t.to_dict()
        d['running_balance'] = round(running, 4)
        rows.append(d)

    tc, td = a.get_totals()
    return jsonify({
        'account': {**a.to_dict(with_balance=False),
                    'total_credits': tc, 'total_debits': td, 'balance': tc - td},
        'transactions': rows,
        'filtered_credits': fc,
        'filtered_debits': fd,
        'filtered_net': fc - fd,
        'count': len(rows),
    })


@app.route('/api/accounts/<int:aid>/txn', methods=['POST'])
def api_txn_create(aid):
    FinAccount.query.get_or_404(aid)
    data = request.get_json() or {}
    txn = FinTransaction(
        account_id=aid,
        txn_date=_parse_date(data.get('txn_date')),
        transaction_date=_parse_date(data.get('transaction_date')),
        credit=float(data.get('credit') or 0),
        debit=float(data.get('debit') or 0),
        credit_remark=(data.get('credit_remark') or '').strip(),
        debit_remark=(data.get('debit_remark') or '').strip(),
        status=data.get('status', 'normal'),
        color_flag=data.get('color_flag', ''),
        tags=data.get('tags', ''),
        priority=int(data.get('priority') or 0),
        notes=(data.get('notes') or '').strip(),
        sort_order=_max_sort(FinTransaction, account_id=aid) + 1,
    )
    db.session.add(txn)
    db.session.flush()
    _audit('txn', txn.id, 'create', after=txn.to_dict())
    db.session.commit()
    return jsonify(txn.to_dict()), 201


@app.route('/api/txn/<int:tid>', methods=['PUT'])
def api_txn_update(tid):
    txn = FinTransaction.query.get_or_404(tid)
    before = txn.to_dict()
    data = request.get_json() or {}

    if 'txn_date' in data:
        txn.txn_date = _parse_date(data['txn_date'])
    if 'transaction_date' in data:
        txn.transaction_date = _parse_date(data['transaction_date'])
    for f in ('credit', 'debit'):
        if f in data:
            setattr(txn, f, float(data[f] or 0))
    for f in ('credit_remark', 'debit_remark', 'status', 'color_flag', 'tags', 'notes'):
        if f in data:
            setattr(txn, f, data[f])
    if 'priority' in data:
        txn.priority = int(data['priority'] or 0)
    if 'sort_order' in data:
        txn.sort_order = int(data['sort_order'])
    txn.updated_at = datetime.utcnow()
    _audit('txn', txn.id, 'update', before=before, after=txn.to_dict())
    db.session.commit()
    return jsonify(txn.to_dict())


@app.route('/api/txn/<int:tid>', methods=['DELETE'])
def api_txn_delete(tid):
    txn = FinTransaction.query.get_or_404(tid)
    aid = txn.account_id
    _audit('txn', tid, 'delete', before=txn.to_dict())
    db.session.delete(txn)
    db.session.commit()
    tc, td = FinAccount.query.get(aid).get_totals()
    return jsonify({'ok': True, 'balance': tc - td, 'total_credits': tc, 'total_debits': td})


@app.route('/api/txn/bulk', methods=['POST'])
def api_txn_bulk():
    data = request.get_json() or {}
    action = data.get('action')
    ids    = data.get('ids', [])
    if not ids:
        return jsonify({'ok': True, 'affected': 0})

    if action == 'delete':
        FinTransaction.query.filter(FinTransaction.id.in_(ids))\
            .delete(synchronize_session=False)
    elif action == 'status':
        FinTransaction.query.filter(FinTransaction.id.in_(ids))\
            .update({'status': data.get('status', 'normal')}, synchronize_session=False)
    elif action == 'color_flag':
        FinTransaction.query.filter(FinTransaction.id.in_(ids))\
            .update({'color_flag': data.get('color_flag', '')}, synchronize_session=False)
    elif action == 'priority':
        FinTransaction.query.filter(FinTransaction.id.in_(ids))\
            .update({'priority': int(data.get('priority', 0))}, synchronize_session=False)

    _audit('txn_bulk', None, action, desc=f'{action} on {len(ids)} txns')
    db.session.commit()
    return jsonify({'ok': True, 'affected': len(ids)})


# ═══════════════════════════════════════════════════════════════
# ROUTES — Search & Autocomplete
# ═══════════════════════════════════════════════════════════════

@app.route('/api/search')
def api_search():
    q = (request.args.get('q') or '').strip()
    if len(q) < 1:
        return jsonify({'results': [], 'total': 0})

    limit = min(int(request.args.get('limit', 150)), 500)
    pat   = f'%{q}%'

    accts = FinAccount.query.filter(
        FinAccount.is_archived == False,
        or_(FinAccount.name.ilike(pat), FinAccount.number.ilike(pat))
    ).limit(20).all()

    txns = db.session.query(FinTransaction, FinAccount)\
        .join(FinAccount, FinTransaction.account_id == FinAccount.id)\
        .filter(
            FinAccount.is_archived == False,
            or_(
                FinTransaction.credit_remark.ilike(pat),
                FinTransaction.debit_remark.ilike(pat),
                FinTransaction.notes.ilike(pat),
                FinTransaction.tags.ilike(pat),
            )
        ).limit(limit).all()

    results = [
        dict(type='account', id=a.id, account_id=a.id,
             title=a.name, subtitle=a.number or 'Account')
        for a in accts
    ] + [
        dict(type='transaction', id=t.id, txn_id=t.id, account_id=acct.id,
             title=t.credit_remark or t.debit_remark or 'Entry',
             subtitle=f'{acct.name}',
             amount=t.credit if t.credit else -t.debit,
             date=t.txn_date.isoformat() if t.txn_date else None,
             color_flag=t.color_flag, status=t.status)
        for t, acct in txns
    ]
    return jsonify({'results': results, 'total': len(results)})


@app.route('/api/autocomplete')
def api_autocomplete():
    q    = (request.args.get('q') or '').strip()
    fld  = request.args.get('field', 'remark')
    if not q:
        return jsonify([])
    pat = f'%{q}%'
    r1 = db.session.query(FinTransaction.credit_remark).filter(
        FinTransaction.credit_remark.ilike(pat), FinTransaction.credit_remark != ''
    ).distinct().limit(8).all()
    r2 = db.session.query(FinTransaction.debit_remark).filter(
        FinTransaction.debit_remark.ilike(pat), FinTransaction.debit_remark != ''
    ).distinct().limit(8).all()
    suggestions = sorted(set(
        [x[0] for x in r1 if x[0]] + [x[0] for x in r2 if x[0]]
    ))[:12]
    return jsonify(suggestions)


@app.route('/api/recent_remarks')
def api_recent_remarks():
    rows = db.session.query(FinTransaction.credit_remark)\
        .filter(FinTransaction.credit_remark != '')\
        .order_by(FinTransaction.id.desc()).limit(30).all()
    rows2 = db.session.query(FinTransaction.debit_remark)\
        .filter(FinTransaction.debit_remark != '')\
        .order_by(FinTransaction.id.desc()).limit(30).all()
    seen, out = set(), []
    for (r,) in rows + rows2:
        if r and r not in seen:
            seen.add(r)
            out.append(r)
        if len(out) >= 20:
            break
    return jsonify(out)


# ═══════════════════════════════════════════════════════════════
# ROUTES — Analytics
# ═══════════════════════════════════════════════════════════════

@app.route('/api/accounts/<int:aid>/analytics')
def api_account_analytics(aid):
    a = FinAccount.query.get_or_404(aid)
    txns = FinTransaction.query.filter_by(account_id=aid)\
        .order_by(FinTransaction.sort_order, FinTransaction.id).all()

    months = defaultdict(lambda: {'credits': 0.0, 'debits': 0.0, 'count': 0})
    for t in txns:
        key = t.txn_date.strftime('%Y-%m') if t.txn_date else 'undated'
        months[key]['credits'] += t.credit
        months[key]['debits']  += t.debit
        months[key]['count']   += 1

    monthly = [{'month': k, **v, 'net': v['credits'] - v['debits']}
               for k, v in sorted(months.items())]

    status_counts = dict(
        db.session.query(FinTransaction.status, func.count())
        .filter_by(account_id=aid).group_by(FinTransaction.status).all()
    )
    tc, td = a.get_totals()
    return jsonify({
        'account': {**a.to_dict(with_balance=False), 'total_credits': tc,
                    'total_debits': td, 'balance': tc - td},
        'monthly': monthly,
        'status_counts': status_counts,
        'total_txns': len(txns),
    })


@app.route('/api/analytics/global')
def api_global_analytics():
    months = defaultdict(lambda: {'credits': 0.0, 'debits': 0.0})
    txns = FinTransaction.query.join(FinAccount)\
        .filter(FinAccount.is_archived == False,
                FinTransaction.txn_date.isnot(None)).all()
    for t in txns:
        key = t.txn_date.strftime('%Y-%m')
        months[key]['credits'] += t.credit
        months[key]['debits']  += t.debit
    return jsonify([{'month': k, **v, 'net': v['credits'] - v['debits']}
                    for k, v in sorted(months.items())])


# ═══════════════════════════════════════════════════════════════
# ROUTES — Audit Log
# ═══════════════════════════════════════════════════════════════

@app.route('/api/audit')
def api_audit():
    limit   = min(int(request.args.get('limit', 200)), 500)
    date_s  = request.args.get('date')   # YYYY-MM-DD  → filter by that day
    since_s = request.args.get('since')  # ISO ts      → entries after this time (session mode)

    q = FinAuditLog.query.order_by(FinAuditLog.id.desc())
    if date_s:
        try:
            d = datetime.strptime(date_s, '%Y-%m-%d').date()
            q = q.filter(func.date(FinAuditLog.created_at) == d)
        except Exception:
            pass
    elif since_s:
        try:
            since_dt = datetime.fromisoformat(since_s.rstrip('Z'))
            q = q.filter(FinAuditLog.created_at >= since_dt)
        except Exception:
            pass

    logs = q.limit(limit).all()

    # Pre-load account names to resolve txn entries without N+1 queries
    acct_map = {a.id: a.name for a in FinAccount.query.all()}

    result = []
    for l in logs:
        row = dict(
            id=l.id, entity_type=l.entity_type, entity_id=l.entity_id,
            action=l.action, description=l.description or '',
            before_data=l.before_data or '', after_data=l.after_data or '',
            created_at=l.created_at.isoformat() if l.created_at else '',
        )
        # Attach account name / id by parsing after_data or before_data
        for src in (l.after_data, l.before_data):
            if src:
                try:
                    d = json.loads(src)
                    aid = d.get('account_id')
                    if aid:
                        row['account_id']   = aid
                        row['account_name'] = acct_map.get(aid, '')
                        break
                except Exception:
                    pass
        result.append(row)
    return jsonify(result)


# ═══════════════════════════════════════════════════════════════
# ROUTES — Export
# ═══════════════════════════════════════════════════════════════

@app.route('/api/export/excel')
def api_export():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Summary'

    # Styles
    H_FILL  = PatternFill('solid', start_color='1E3A5F')
    H_FONT  = Font(color='FFFFFF', bold=True, name='Arial', size=10)
    NUM_FMT = '#,##0.00'
    border  = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    kpis = compute_kpis()
    ws['A1'] = 'Financial Account Management System — Export'
    ws['A1'].font = Font(bold=True, size=14)
    ws['C1'] = datetime.now().strftime('%Y-%m-%d %H:%M')

    ws.append([])
    for label, val in [
        ('Manual Balance (H + T + Dolph)', kpis['manual_balance']),
        ('System Calculated Total',        kpis['system_total']),
        ('Difference',                      kpis['difference']),
        ('T Value (manual)',                kpis['t_value']),
        ('H Accounts Balance',              kpis['h_balance']),
        ('Dolph Accounts Balance',          kpis['dolph_balance']),
    ]:
        row = [label, val]
        ws.append(row)
        ws.cell(ws.max_row, 2).number_format = NUM_FMT

    ws.column_dimensions['A'].width = 36
    ws.column_dimensions['B'].width = 18

    accounts = FinAccount.query.filter_by(is_archived=False)\
        .order_by(FinAccount.sort_order, FinAccount.id).all()

    for acct in accounts:
        safe = acct.name[:28].translate(str.maketrans('/\\[]?*:', '-------'))
        wsa = wb.create_sheet(title=safe or f'Acct_{acct.id}')

        # Account header
        wsa.append([acct.name, acct.number, '', f'Group: {acct.group_type}'])
        wsa['A1'].font = Font(bold=True, size=12)
        wsa.append([])

        headers = ['Date', 'Credit', 'Credit Remark', 'Debit', 'Debit Remark',
                   'Status', 'Tags', 'Running Balance']
        wsa.append(headers)
        for col_i, _ in enumerate(headers, 1):
            cell = wsa.cell(3, col_i)
            cell.fill = H_FILL
            cell.font = H_FONT
            cell.alignment = Alignment(horizontal='center')

        running = 0.0
        txns = FinTransaction.query.filter_by(account_id=acct.id)\
            .order_by(FinTransaction.sort_order, FinTransaction.id).all()

        for t in txns:
            running += t.credit - t.debit
            wsa.append([
                t.txn_date.isoformat() if t.txn_date else '',
                t.credit or None, t.credit_remark,
                t.debit or None, t.debit_remark,
                t.status, t.tags, running,
            ])
            r = wsa.max_row
            for c in (2, 4, 8):
                wsa.cell(r, c).number_format = NUM_FMT

        # Totals row
        if txns:
            n = len(txns)
            data_start = 4
            data_end = data_start + n - 1
            wsa.append([
                'TOTAL',
                f'=SUM(B{data_start}:B{data_end})', '',
                f'=SUM(D{data_start}:D{data_end})', '',
                '', '', running
            ])
            tr = wsa.max_row
            for c in (2, 4, 8):
                cell = wsa.cell(tr, c)
                cell.font = Font(bold=True)
                cell.number_format = NUM_FMT

        for col_i, width in enumerate([14, 14, 34, 14, 34, 12, 24, 16], 1):
            wsa.column_dimensions[get_column_letter(col_i)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f'fin_accounts_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(buf, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ═══════════════════════════════════════════════════════════════
# ROUTES — Import Excel
# ═══════════════════════════════════════════════════════════════

@app.route('/api/import/excel', methods=['POST'])
def api_import_excel():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    f = request.files['file']
    if not f.filename.lower().endswith('.xlsx'):
        return jsonify({'error': 'File must be .xlsx'}), 400

    try:
        wb = load_workbook(f, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        return jsonify({'error': str(e)}), 400

    # Determine if file has standard export format (Account / Date / Credit / Debit)
    # Look for a header row
    header_row = None
    for i, row in enumerate(rows[:5]):
        cells = [str(c).lower().strip() if c else '' for c in row]
        if 'credit' in cells and 'debit' in cells:
            header_row = i
            break

    imported_txns = 0
    if header_row is not None:
        headers = [str(c).lower().strip() if c else '' for c in rows[header_row]]

        def col(name):
            try:
                return headers.index(name)
            except ValueError:
                return None

        c_date   = col('date')
        c_credit = col('credit')
        c_debit  = col('debit')
        c_cr_rem = col('credit remark')
        c_dr_rem = col('debit remark')
        c_status = col('status')

        for row in rows[header_row + 1:]:
            if not any(row):
                continue
            credit = float(row[c_credit] or 0) if c_credit is not None else 0
            debit  = float(row[c_debit]  or 0) if c_debit  is not None else 0
            if credit == 0 and debit == 0:
                continue
            txn = FinTransaction(
                account_id=1,  # caller should pass account_id in query param
                txn_date=_parse_date(row[c_date]) if c_date is not None else None,
                credit=credit, debit=debit,
                credit_remark=str(row[c_cr_rem] or '') if c_cr_rem is not None else '',
                debit_remark=str(row[c_dr_rem] or '') if c_dr_rem is not None else '',
                status=str(row[c_status] or 'normal') if c_status is not None else 'normal',
                sort_order=_max_sort(FinTransaction, account_id=1) + 1,
            )
            db.session.add(txn)
            imported_txns += 1

    db.session.commit()
    return jsonify({'ok': True, 'imported_transactions': imported_txns})


# ═══════════════════════════════════════════════════════════════
# STARTUP
# ═══════════════════════════════════════════════════════════════

def init_db():
    with app.app_context():
        db.create_all()
        # Add transaction_date column if it doesn't exist yet (migration)
        try:
            db.session.execute(text('ALTER TABLE fin_transactions ADD COLUMN transaction_date DATE'))
            db.session.commit()
        except Exception:
            db.session.rollback()
        # Default settings
        defaults = {'t_value': '0', 'previous_value': '', 'updated_till': '', 'currency_symbol': '₹', 'app_name': 'FinLedger Pro'}
        for k, v in defaults.items():
            if not FinSetting.query.get(k):
                db.session.add(FinSetting(key=k, value=v))
        db.session.commit()


# Auto-initialise DB when imported by gunicorn / any WSGI server
init_db()

if __name__ == '__main__':
    init_db()
    _port = int(os.environ.get('FIN_PORT', 5051))
    print(f'FinLedger Pro → http://127.0.0.1:{_port}')
    app.run(host='0.0.0.0', port=_port, debug=True)

